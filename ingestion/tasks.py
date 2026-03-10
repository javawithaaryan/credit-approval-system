import os
from datetime import datetime, date

from celery import shared_task
from django.conf import settings
import logging

logger = logging.getLogger(__name__)


def to_int(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        try:
            return int(float(val.strip()))
        except (ValueError, TypeError):
            return 0
    return 0


def to_date(val):
    if val is None:
        return None
    if isinstance(val, date):
        # datetime.datetime is a subclass of datetime.date, so this works for both
        if hasattr(val, 'date'):
            return val.date()
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None
    return None


@shared_task(name='ingestion.tasks.ingest_data')
def ingest_data():
    """
    Ingest customer and loan data from Excel files into the database.

    Only runs if the Customer and Loan tables are empty to prevent
    duplicate ingestion. Uses bulk_create for efficiency and robustness.
    """
    from customers.models import Customer
    from loans.models import Loan
    import openpyxl

    data_dir = getattr(settings, 'DATA_DIR', os.path.join(settings.BASE_DIR, 'data'))

    # --- Ingest Customers ---
    if not Customer.objects.exists():
        customer_file = os.path.join(data_dir, 'customer_data.xlsx')
        if os.path.exists(customer_file):
            logger.info('Ingesting customer data from %s', customer_file)
            wb = openpyxl.load_workbook(customer_file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            customers = []
            for row in rows:
                if row[0] is None:
                    continue
                try:
                    customers.append(Customer(
                        customer_id=to_int(row[0]),
                        first_name=str(row[1]).strip() if row[1] else "",
                        last_name=str(row[2]).strip() if row[2] else "",
                        age=to_int(row[3]),
                        phone_number=to_int(row[4]),
                        monthly_salary=to_int(row[5]),
                        approved_limit=to_int(row[6]),
                        current_debt=0.0,
                    ))
                except Exception as e:
                    logger.error("Error parsing customer row %s: %s", row, e)
            Customer.objects.bulk_create(customers, batch_size=1000)
            wb.close()
            logger.info('Successfully ingested %d customers', len(customers))
        else:
            logger.warning('Customer data file not found: %s', customer_file)
    else:
        logger.info('Customer table already has data, skipping ingestion.')

    # --- Ingest Loans ---
    if not Loan.objects.exists():
        loan_file = os.path.join(data_dir, 'loan_data.xlsx')
        if os.path.exists(loan_file):
            logger.info('Ingesting loan data from %s', loan_file)
            wb = openpyxl.load_workbook(loan_file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            loans = []
            for row in rows:
                if row[0] is None:
                    continue
                try:
                    customer_id = to_int(row[0])
                    loan_id = to_int(row[1])
                    loan_amount = float(row[2]) if row[2] is not None else 0.0
                    tenure = to_int(row[3])
                    interest_rate = float(row[4]) if row[4] is not None else 0.0
                    monthly_repayment = float(row[5]) if row[5] is not None else 0.0
                    emis_paid_on_time = to_int(row[6])
                    start_date = to_date(row[7])
                    end_date = to_date(row[8])

                    if not start_date or not end_date:
                        continue

                    loans.append(Loan(
                        customer_id=customer_id,
                        loan_id=loan_id,
                        loan_amount=loan_amount,
                        tenure=tenure,
                        interest_rate=interest_rate,
                        monthly_repayment=monthly_repayment,
                        emis_paid_on_time=emis_paid_on_time,
                        start_date=start_date,
                        end_date=end_date,
                    ))
                except Exception as e:
                    logger.error("Error parsing loan row %s: %s", row, e)
            Loan.objects.bulk_create(loans, batch_size=1000)
            wb.close()
            logger.info('Successfully ingested %d loans', len(loans))
        else:
            logger.warning('Loan data file not found: %s', loan_file)
    else:
        logger.info('Loan table already has data, skipping ingestion.')

    return 'Data ingestion complete'
